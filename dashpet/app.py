from flask import Flask, jsonify, request, send_from_directory
import openpyxl
from datetime import datetime, date
import os

app = Flask(__name__, static_folder='.')
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exames-controle.xlsx')

DATA_ROW_START = 4  # Row index (1-based) where data starts in Exames/Gastos sheets


def _clean(val):
    if val is None:
        return ''
    if isinstance(val, str) and ('IFERROR' in val or val.startswith('=')):
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    return val


def load_wb():
    return openpyxl.load_workbook(XLSX_PATH, data_only=True)


def load_wb_formulas():
    return openpyxl.load_workbook(XLSX_PATH)


def save_wb(wb):
    wb.save(XLSX_PATH)


# ── Parâmetros ──────────────────────────────────────────────────────────────
def get_params():
    wb = load_wb()
    ws = wb['Parâmetros']
    rows = list(ws.iter_rows(values_only=True))

    cats = {
        'Ultrassonografia': [],
        'Cardiologia': [],
        'RaioX': [],
        'Tomografia': [],
        'Outros': [],
    }
    cat_cols = {
        'Ultrassonografia': 1,
        'Cardiologia': 3,
        'RaioX': 5,
        'Tomografia': 7,
        'Outros': 9,
    }
    payment = []

    for row in rows[3:]:  # skip headers
        for cat, col in cat_cols.items():
            v = row[col] if col < len(row) else None
            if v and not isinstance(v, str) or (isinstance(v, str) and v.strip() and not v.startswith('=')):
                if v:
                    cats[cat].append(str(v).strip())
        pv = row[11] if len(row) > 11 else None
        if pv and isinstance(pv, str) and pv.strip():
            payment.append(pv.strip())

    return {'exam_categories': cats, 'payment_methods': payment}


# ── Exames ───────────────────────────────────────────────────────────────────
def get_exames():
    wb = load_wb()
    ws = wb['Exames']
    exames = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DATA_ROW_START:
            continue
        # Columns (0-based): 1=Data, 2=Vet, 3=Clínica, 4=Animal, 5=Tipo, 6=Específico, 7=Valor, 8=Pagamento
        data = _clean(row[1] if len(row) > 1 else None)
        tipo = _clean(row[5] if len(row) > 5 else None)
        valor = row[7] if len(row) > 7 else None
        if isinstance(valor, str) and ('IFERROR' in valor or valor.startswith('=')):
            valor = None
        if not data and not tipo and not valor:
            continue
        exames.append({
            'row': i,
            'data': data,
            'vet': _clean(row[2] if len(row) > 2 else None),
            'clinica': _clean(row[3] if len(row) > 3 else None),
            'animal': _clean(row[4] if len(row) > 4 else None),
            'tipo_exame': tipo,
            'exame_especifico': _clean(row[6] if len(row) > 6 else None),
            'valor': float(valor) if valor and str(valor).replace('.', '', 1).isdigit() else (float(valor) if isinstance(valor, (int, float)) else 0),
            'forma_pagamento': _clean(row[8] if len(row) > 8 else None),
        })
    return exames


def add_exame(data):
    wb = load_wb_formulas()
    ws = wb['Exames']
    # Find first empty row (B column is None)
    target_row = None
    for i, row in enumerate(ws.iter_rows(min_row=DATA_ROW_START, values_only=False), start=DATA_ROW_START):
        if row[1].value is None:  # col B
            target_row = row[1].row
            break
    if target_row is None:
        target_row = ws.max_row + 1

    try:
        date_val = datetime.strptime(data['data'], '%Y-%m-%d').date() if data.get('data') else None
    except ValueError:
        date_val = None

    ws.cell(row=target_row, column=2).value = date_val
    ws.cell(row=target_row, column=3).value = data.get('vet', '')
    ws.cell(row=target_row, column=4).value = data.get('clinica', '')
    ws.cell(row=target_row, column=5).value = data.get('animal', '')
    ws.cell(row=target_row, column=6).value = data.get('tipo_exame', '')
    ws.cell(row=target_row, column=7).value = data.get('exame_especifico', '')
    try:
        ws.cell(row=target_row, column=8).value = float(data.get('valor', 0))
    except (ValueError, TypeError):
        ws.cell(row=target_row, column=8).value = 0
    ws.cell(row=target_row, column=9).value = data.get('forma_pagamento', '')
    save_wb(wb)
    return target_row


def delete_exame(row_num):
    wb = load_wb_formulas()
    ws = wb['Exames']
    # Clear the data columns B-I
    for col in range(2, 10):
        ws.cell(row=row_num, column=col).value = None
    save_wb(wb)


# ── Gastos ────────────────────────────────────────────────────────────────────
def get_gastos():
    wb = load_wb()
    ws = wb['Gastos']
    gastos = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DATA_ROW_START:
            continue
        data = _clean(row[1] if len(row) > 1 else None)
        tipo = _clean(row[2] if len(row) > 2 else None)
        descricao = _clean(row[3] if len(row) > 3 else None)
        valor = row[7] if len(row) > 7 else None
        if isinstance(valor, str) and ('IFERROR' in valor or valor.startswith('=')):
            valor = None
        if not data and not tipo and not descricao:
            continue
        try:
            valor_f = float(valor) if valor is not None else 0
        except (ValueError, TypeError):
            valor_f = 0
        gastos.append({
            'row': i,
            'data': data,
            'tipo': tipo,
            'descricao': descricao,
            'valor': valor_f,
        })
    return gastos


def add_gasto(data):
    wb = load_wb_formulas()
    ws = wb['Gastos']
    target_row = None
    for row in ws.iter_rows(min_row=DATA_ROW_START, values_only=False):
        if row[1].value is None:
            target_row = row[1].row
            break
    if target_row is None:
        target_row = ws.max_row + 1

    try:
        date_val = datetime.strptime(data['data'], '%Y-%m-%d').date() if data.get('data') else None
    except ValueError:
        date_val = None

    ws.cell(row=target_row, column=2).value = date_val
    ws.cell(row=target_row, column=3).value = data.get('tipo', 'Despesa')
    ws.cell(row=target_row, column=4).value = data.get('descricao', '')
    try:
        ws.cell(row=target_row, column=8).value = float(data.get('valor', 0))
    except (ValueError, TypeError):
        ws.cell(row=target_row, column=8).value = 0
    save_wb(wb)
    return target_row


def delete_gasto(row_num):
    wb = load_wb_formulas()
    ws = wb['Gastos']
    for col in range(2, 9):
        ws.cell(row=row_num, column=col).value = None
    save_wb(wb)


# ── Summary ───────────────────────────────────────────────────────────────────
def get_summary():
    exames = get_exames()
    gastos = get_gastos()

    total_receita = sum(e['valor'] for e in exames if e['forma_pagamento'] != 'Em Aberto')
    em_aberto = sum(e['valor'] for e in exames if e['forma_pagamento'] == 'Em Aberto')
    total_despesa = sum(g['valor'] for g in gastos if g['tipo'] == 'Despesa')
    total_investimento = sum(g['valor'] for g in gastos if g['tipo'] == 'Investimento')

    # Monthly breakdown
    monthly = {}
    for e in exames:
        if e['data'] and len(e['data']) >= 7:
            month = e['data'][:7]
            if month not in monthly:
                monthly[month] = {'receita': 0, 'despesa': 0, 'exames': 0}
            monthly[month]['receita'] += e['valor'] if e['forma_pagamento'] != 'Em Aberto' else 0
            monthly[month]['exames'] += 1
    for g in gastos:
        if g['data'] and len(g['data']) >= 7:
            month = g['data'][:7]
            if month not in monthly:
                monthly[month] = {'receita': 0, 'despesa': 0, 'exames': 0}
            if g['tipo'] == 'Despesa':
                monthly[month]['despesa'] += g['valor']

    # Exam type breakdown
    by_type = {}
    for e in exames:
        t = e['tipo_exame'] or 'Sem tipo'
        by_type[t] = by_type.get(t, 0) + 1

    return {
        'total_exames': len(exames),
        'total_receita': total_receita,
        'em_aberto': em_aberto,
        'total_despesa': total_despesa,
        'total_investimento': total_investimento,
        'lucro': total_receita - total_despesa,
        'monthly': dict(sorted(monthly.items())),
        'by_type': by_type,
    }


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/api/params')
def api_params():
    try:
        return jsonify(get_params())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/exames', methods=['GET'])
def api_get_exames():
    try:
        return jsonify(get_exames())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/exames', methods=['POST'])
def api_add_exame():
    try:
        data = request.get_json()
        row = add_exame(data)
        return jsonify({'ok': True, 'row': row})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/exames/<int:row_num>', methods=['DELETE'])
def api_delete_exame(row_num):
    try:
        delete_exame(row_num)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/gastos', methods=['GET'])
def api_get_gastos():
    try:
        return jsonify(get_gastos())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/gastos', methods=['POST'])
def api_add_gasto():
    try:
        data = request.get_json()
        row = add_gasto(data)
        return jsonify({'ok': True, 'row': row})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/gastos/<int:row_num>', methods=['DELETE'])
def api_delete_gasto(row_num):
    try:
        delete_gasto(row_num)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/summary')
def api_summary():
    try:
        return jsonify(get_summary())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
